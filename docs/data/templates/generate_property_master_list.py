#!/usr/bin/env python3
"""
Property Master List Template Generator
‡∏™‡∏£‡πâ‡∏≤‡∏á Excel template ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏à‡∏±‡∏î‡∏Å‡∏≤‡∏£‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• property ‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î‡∏Ç‡∏≠‡∏á AMP

This script generates the Property_Master_List.xlsx file according to the 
specifications defined in PROPERTY_MASTER_LIST.md

# Dependencies:
#   This script requires the `openpyxl` package to generate Excel files.
#   Install it before running this script, for example:
#       pip install openpyxl
#   or, if using Poetry:
#       poetry add --group docs openpyxl
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.workbook import Workbook as WorkbookType
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:
    raise SystemExit(
        "This script requires the 'openpyxl' package. "
        "Install it with: pip install openpyxl"
    ) from exc


def create_property_master_list() -> str:
    """Create the Property Master List Excel template"""
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create all tabs
    ws_all = wb.create_sheet("01_All_Properties")
    ws_projects = wb.create_sheet("02_Projects_New")
    ws_resale = wb.create_sheet("03_Resale_Secondary")
    ws_rental_lt = wb.create_sheet("04_Rental_Long_Term")
    ws_rental_st = wb.create_sheet("05_Rental_Short_Term")
    ws_sold = wb.create_sheet("06_Sold_Rented")
    ws_pending = wb.create_sheet("07_Pending")
    ws_readme = wb.create_sheet("README")
    
    # Define column headers for main tab (01_All_Properties)
    headers = [
        ("A", "Property_ID", "Text", "Unique identifier", "PROP-2026-001"),
        ("B", "Status", "Dropdown", "Current status", "Available"),
        ("C", "Category", "Dropdown", "Property category", "Project"),
        ("D", "Type", "Dropdown", "Property type", "Condo"),
        ("E", "Project_Name", "Text", "Project/Building name", "The Riviera Jomtien"),
        ("F", "Unit_Number", "Text", "Unit number", "A205"),
        ("G", "Location_Area", "Dropdown", "General area", "Jomtien"),
        ("H", "Location_Sub", "Text", "Specific location", "Jomtien Beach Road"),
        ("I", "Address", "Text", "Full address", "123 Jomtien Beach Rd"),
        ("J", "Latitude", "Number", "GPS latitude", "12.XXXXX"),
        ("K", "Longitude", "Number", "GPS longitude", "100.XXXXX"),
        ("L", "Price_Sale", "Number", "Sale price (THB)", "2500000"),
        ("M", "Price_Rent", "Number", "Monthly rent (THB)", "15000"),
        ("N", "Currency", "Dropdown", "Currency code", "THB"),
        ("O", "Price_Per_SQM", "Formula", "Price per square meter", "62500"),
        ("P", "Negotiable", "Dropdown", "Price negotiable?", "Yes"),
        ("Q", "Commission_Rate", "Percent", "Commission %", "3%"),
        ("R", "Owner_Type", "Dropdown", "Owner type", "Freehold"),
        ("S", "Foreign_Quota", "Dropdown", "Foreign ownership?", "Yes"),
        ("T", "Bedrooms", "Number", "Number of bedrooms", "1"),
        ("U", "Bathrooms", "Number", "Number of bathrooms", "1"),
        ("V", "Size_SQM", "Number", "Size in square meters", "40"),
        ("W", "Size_SQW", "Formula", "Size in square wah", "10"),
        ("X", "Floor_Level", "Text", "Floor number", "12"),
        ("Y", "View", "Text", "View description", "Sea view"),
        ("Z", "Furnishing", "Dropdown", "Furnishing status", "Fully Furnished"),
        ("AA", "Parking", "Number", "Parking spaces", "1"),
        ("AB", "Facilities", "Text", "Key facilities", "Pool, Gym, 24h Security"),
        ("AC", "Condition", "Dropdown", "Property condition", "Excellent"),
        ("AD", "Year_Built", "Number", "Year of construction", "2023"),
        ("AE", "Year_Renovated", "Number", "Last renovation year", "2025"),
        ("AF", "Title_Deed", "Text", "Title deed number", "12345/6789"),
        ("AG", "Developer", "Text", "Developer name", "ABC Development"),
        ("AH", "Photos_Link", "URL", "Google Drive photos folder", "[URL]"),
        ("AI", "Video_Link", "URL", "Property video link", "[URL]"),
        ("AJ", "Brochure_Link", "URL", "Brochure PDF link", "[URL]"),
        ("AK", "Virtual_Tour", "URL", "360¬∞ tour link", "[URL]"),
        ("AL", "Listing_URL", "URL", "Website listing page", "[URL]"),
        ("AM", "Source", "Dropdown", "How we got this listing", "Owner Direct"),
        ("AN", "Source_Contact", "Text", "Contact at source", "John Doe, 0XX-XXX-XXXX"),
        ("AO", "Assigned_Agent", "Dropdown", "Agent responsible", "Somchai S."),
        ("AP", "Priority", "Dropdown", "Marketing priority", "High"),
        ("AQ", "Featured", "Dropdown", "Feature on website?", "Yes"),
        ("AR", "Active_Marketing", "Dropdown", "Currently marketing?", "Yes"),
        ("AS", "Date_Added", "Date", "Date added to system", "2026-01-26"),
        ("AT", "Date_Updated", "Date", "Last update date", "2026-01-26"),
        ("AU", "Date_Available", "Date", "Available from date", "2026-02-01"),
        ("AV", "Views_Count", "Number", "Number of online views", "15"),
        ("AW", "Viewings_Count", "Number", "Physical viewings done", "3"),
        ("AX", "Notes_Internal", "Text", "Internal notes", "Good deal, motivated seller"),
        ("AY", "Tags", "Text", "Search tags", "sea view, investment, new"),
        ("AZ", "Language", "Text", "Listing languages", "TH, EN"),
        ("BA", "WhatsApp_Message", "Text", "Pre-formatted message", "Hello, I'm interested in Property PROP-2026-001"),
        ("BB", "QR_Code_URL", "URL", "Generated wa.me link", "https://wa.me/66XXX?text=..."),
        ("BC", "Print_Status", "Dropdown", "Currently printed?", "Yes"),
        ("BD", "Print_Slot", "Number", "Slot number if printed", "5"),
        ("BE", "FB_Posted_Date", "Date", "Last Facebook post date", "2026-01-26"),
        ("BF", "FB_Post_Link", "URL", "Link to Facebook post", "[URL]"),
        ("BG", "Website_Listed", "Dropdown", "Listed on website?", "Yes"),
        ("BH", "Website_URL", "URL", "Property page URL", "[URL]"),
        ("BI", "Marketing_Priority", "Dropdown", "Marketing priority level", "High"),
    ]
    
    # Setup main tab
    setup_main_tab(ws_all, headers)
    
    # Setup other tabs (Projects, Resale, etc.) with their headers
    setup_projects_tab(ws_projects, headers)
    setup_resale_tab(ws_resale, headers)
    setup_rental_lt_tab(ws_rental_lt, headers)
    setup_rental_st_tab(ws_rental_st, headers)
    setup_sold_rented_tab(ws_sold, headers)
    setup_pending_tab(ws_pending, headers)
    setup_readme_tab(ws_readme)
    
    # Save the workbook
    filename = "Property_Master_List.xlsx"
    wb.save(filename)
    print(f"‚úÖ Created: {filename}")
    return filename


def setup_main_tab(ws: Worksheet, headers: list) -> None:
    """Setup the main 01_All_Properties tab with all columns and validations"""
    
    # Set up headers
    for idx, (col, name, dtype, desc, example) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Set column widths
    column_widths = {
        'A': 15, 'B': 12, 'C': 12, 'D': 12, 'E': 25, 'F': 12, 'G': 15, 'H': 20,
        'I': 30, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 10, 'O': 12, 'P': 12,
        'Q': 12, 'R': 12, 'S': 12, 'T': 10, 'U': 10, 'V': 10, 'W': 10, 'X': 10,
        'Y': 15, 'Z': 15, 'AA': 10, 'AB': 30, 'AC': 15, 'AD': 12, 'AE': 12,
        'AF': 15, 'AG': 20, 'AH': 15, 'AI': 15, 'AJ': 15, 'AK': 15, 'AL': 15,
        'AM': 15, 'AN': 25, 'AO': 15, 'AP': 12, 'AQ': 12, 'AR': 12, 'AS': 12,
        'AT': 12, 'AU': 12, 'AV': 12, 'AW': 12, 'AX': 30, 'AY': 20, 'AZ': 12,
        'BA': 40, 'BB': 40, 'BC': 12, 'BD': 10, 'BE': 12, 'BF': 15, 'BG': 12,
        'BH': 15, 'BI': 15
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add formulas to row 2 as examples
    ws['A2'] = '=IF($AS2<>"", "PROP-"&TEXT(YEAR($AS2),"0000")&"-"&TEXT(ROW()-1,"000"), "")'
    ws['O2'] = '=IF(AND(L2>0, V2>0), L2/V2, "N/A")'
    ws['W2'] = '=IF(V2>0, V2/4, "")'
    ws['BA2'] = '="Hello, I\'m interested in Property "&A2'
    ws['BB2'] = '="https://wa.me/66XXXXXXXXX?text="&ENCODEURL("Hello, I\'m interested in Property "&A2)'
    
    # Create data validation lists
    validations = {
        'Status': ['Available', 'Reserved', 'Sold', 'Rented', 'Pending', 'Off Market'],
        'Category': ['Project', 'Resale', 'Rental-LT', 'Rental-ST'],
        'Type': ['Condo', 'Villa', 'House', 'Townhouse', 'Land', 'Commercial'],
        'Location_Area': ['Pattaya City', 'Jomtien', 'Na Jomtien', 'Pratumnak', 'Wongamat', 
                          'Bang Saray', 'Huay Yai', 'East Pattaya', 'Other'],
        'Currency': ['THB', 'USD', 'EUR'],
        'Owner_Type': ['Freehold', 'Leasehold', 'Company'],
        'Furnishing': ['Unfurnished', 'Partly Furnished', 'Fully Furnished'],
        'Condition': ['New', 'Excellent', 'Good', 'Fair', 'Needs Renovation'],
        'Source': ['Owner Direct', 'Agent', 'LINE Group', 'Facebook', 'Website', 'Walk-in', 'Referral'],
        'Priority': ['Low', 'Medium', 'High'],
        'Yes_No': ['Yes', 'No', 'N/A'],
        'Yes_No_Simple': ['Yes', 'No'],
    }
    
    # Apply data validation
    validation_columns = {
        'B': 'Status',
        'C': 'Category', 
        'D': 'Type',
        'G': 'Location_Area',
        'N': 'Currency',
        'P': 'Yes_No_Simple',
        'R': 'Owner_Type',
        'S': 'Yes_No',
        'Z': 'Furnishing',
        'AC': 'Condition',
        'AM': 'Source',
        'AP': 'Priority',
        'AQ': 'Yes_No_Simple',
        'AR': 'Yes_No_Simple',
        'BC': 'Yes_No_Simple',
        'BG': 'Yes_No_Simple',
        'BI': 'Priority',
    }
    
    for col, val_type in validation_columns.items():
        dv = DataValidation(
            type="list", 
            formula1=f'"{",".join(validations[val_type])}"', 
            allow_blank=True
        )
        dv.error = 'Invalid value'
        dv.errorTitle = 'Invalid Entry'
        ws.add_data_validation(dv)
        # Apply to first 1000 rows
        dv.add(f'{col}2:{col}1000')
    
    # Freeze panes (freeze first row and first column)
    ws.freeze_panes = 'B2'
    
    # Add a sample data row (row 2 with some example values)
    sample_data = {
        'B2': 'Available',
        'C2': 'Project',
        'D2': 'Condo',
        'E2': 'The Riviera Jomtien',
        'F2': 'A1205',
        'G2': 'Jomtien',
        'H2': 'Jomtien Beach Road',
        'L2': 2500000,
        'M2': '',
        'N2': 'THB',
        'P2': 'Yes',
        'R2': 'Freehold',
        'S2': 'Yes',
        'T2': 1,
        'U2': 1,
        'V2': 40,
        'Y2': 'Sea view',
        'Z2': 'Fully Furnished',
        'AA2': 1,
        'AB2': 'Pool, Gym, 24h Security',
        'AC2': 'Excellent',
        'AD2': 2023,
        'AG2': 'ABC Development',
        'AH2': 'https://drive.google.com/...',
        'AM2': 'Owner Direct',
        'AO2': 'Somchai S.',
        'AP2': 'High',
        'AQ2': 'Yes',
        'AR2': 'Yes',
        'AS2': '2026-01-26',
        'AT2': '2026-01-26',
        'AV2': 15,
        'AW2': 3,
        'AX2': 'Good deal, motivated seller',
        'AY2': 'sea view, investment, new',
        'AZ2': 'TH, EN',
        'BC2': 'No',
        'BG2': 'Yes',
        'BI2': 'High',
    }
    
    for cell_ref, value in sample_data.items():
        ws[cell_ref] = value


def setup_projects_tab(ws: Worksheet, base_headers: list) -> None:
    """Setup 02_Projects_New tab with additional columns"""
    # Copy base headers
    for idx, (col, name, dtype, desc, example) in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add additional columns for Projects
    additional_cols = [
        "Total_Units",
        "Units_Available",
        "Transfer_Date",
        "Payment_Plan",
        "Installment_Available"
    ]
    
    start_col = len(base_headers) + 1
    for idx, name in enumerate(additional_cols, start_col):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.freeze_panes = 'B2'


def setup_resale_tab(ws: Worksheet, base_headers: list) -> None:
    """Setup 03_Resale_Secondary tab with additional columns"""
    for idx, (col, name, dtype, desc, example) in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    additional_cols = [
        "Owner_Contact",
        "Reason_Selling",
        "Occupancy",
        "Flexibility"
    ]
    
    start_col = len(base_headers) + 1
    for idx, name in enumerate(additional_cols, start_col):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.freeze_panes = 'B2'


def setup_rental_lt_tab(ws: Worksheet, base_headers: list) -> None:
    """Setup 04_Rental_Long_Term tab with additional columns"""
    for idx, (col, name, dtype, desc, example) in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    additional_cols = [
        "Min_Contract",
        "Utilities_Included",
        "Pets_Allowed",
        "Available_From"
    ]
    
    start_col = len(base_headers) + 1
    for idx, name in enumerate(additional_cols, start_col):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.freeze_panes = 'B2'


def setup_rental_st_tab(ws: Worksheet, base_headers: list) -> None:
    """Setup 05_Rental_Short_Term tab with additional columns"""
    for idx, (col, name, dtype, desc, example) in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    additional_cols = [
        "Min_Nights",
        "Price_High_Season",
        "Price_Low_Season",
        "Cleaning_Fee",
        "Booking_Platforms"
    ]
    
    start_col = len(base_headers) + 1
    for idx, name in enumerate(additional_cols, start_col):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.freeze_panes = 'B2'


def setup_sold_rented_tab(ws: Worksheet, base_headers: list) -> None:
    """Setup 06_Sold_Rented archive tab"""
    for idx, (col, name, dtype, desc, example) in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="666666", end_color="666666", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.freeze_panes = 'B2'


def setup_pending_tab(ws: Worksheet, base_headers: list) -> None:
    """Setup 07_Pending tab"""
    for idx, (col, name, dtype, desc, example) in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=idx)
        cell.value = name
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.freeze_panes = 'B2'


def setup_readme_tab(ws: Worksheet) -> None:
    """Setup README tab with instructions"""
    ws['A1'] = 'PROPERTY MASTER LIST - README'
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    
    readme_content = [
        '',
        'Welcome to the AMP Property Master List!',
        '',
        'This spreadsheet is the central database for all property listings.',
        '',
        'üìã TAB STRUCTURE:',
        '',
        '‚Ä¢ 01_All_Properties - Main comprehensive list of all properties',
        '‚Ä¢ 02_Projects_New - New development projects only',
        '‚Ä¢ 03_Resale_Secondary - Secondary market properties',
        '‚Ä¢ 04_Rental_Long_Term - Long-term rental properties',
        '‚Ä¢ 05_Rental_Short_Term - Short-term/vacation rentals',
        '‚Ä¢ 06_Sold_Rented - Archive of completed transactions',
        '‚Ä¢ 07_Pending - Reserved or in-process properties',
        '‚Ä¢ README - This instructions tab',
        '',
        'üîß GETTING STARTED:',
        '',
        '1. Always work in the 01_All_Properties tab for adding new properties',
        '2. Fill all required fields (marked with ‚úÖ in the schema)',
        '3. Use dropdowns - do not type values directly',
        '4. Property_ID will auto-generate when you set Date_Added',
        '5. Add Google Drive link for property photos',
        '6. Assign to an agent',
        '7. Set Status to "Available"',
        '',
        'üìù REQUIRED FIELDS:',
        '',
        '‚Ä¢ Property_ID (auto-generated)',
        '‚Ä¢ Status, Category, Type',
        '‚Ä¢ Project_Name',
        '‚Ä¢ Location_Area',
        '‚Ä¢ Bedrooms, Bathrooms, Size_SQM',
        '‚Ä¢ Photos_Link',
        '‚Ä¢ Assigned_Agent',
        '‚Ä¢ Active_Marketing',
        '‚Ä¢ Date_Added, Date_Updated',
        '',
        'üîç TIPS:',
        '',
        '‚Ä¢ Use Ctrl+F to search for properties',
        '‚Ä¢ Use filters on column headers for quick filtering',
        '‚Ä¢ Update Date_Updated whenever you modify a property',
        '‚Ä¢ Add internal notes in Notes_Internal column',
        '‚Ä¢ Review properties older than 90 days regularly',
        '',
        'üìä DATA QUALITY:',
        '',
        '‚Ä¢ Keep all property IDs unique',
        '‚Ä¢ Verify photo links are accessible',
        '‚Ä¢ Double-check prices for accuracy',
        '‚Ä¢ Use consistent date format (YYYY-MM-DD)',
        '‚Ä¢ Keep status up to date',
        '',
        'üí° For detailed documentation, see:',
        'docs/data/templates/PROPERTY_MASTER_LIST.md',
        '',
        '---',
        'Created: 2026-01-26',
        'Version: 1.0',
    ]
    
    for idx, line in enumerate(readme_content, 2):
        ws[f'A{idx}'] = line
        if line.startswith('‚Ä¢'):
            ws[f'A{idx}'].font = Font(size=10)
        elif line.startswith(('üìã', 'üîß', 'üìù', 'üîç', 'üìä', 'üí°')):
            ws[f'A{idx}'].font = Font(bold=True, size=12, color="366092")
    
    ws.column_dimensions['A'].width = 80


if __name__ == "__main__":
    create_property_master_list()
