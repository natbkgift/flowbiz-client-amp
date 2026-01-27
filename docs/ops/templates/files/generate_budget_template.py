#!/usr/bin/env python3
"""
Budget Calculator Template Generator
สร้าง Excel template สำหรับวางแผนงบประมาณโฆษณาอสังหาริมทรัพย์
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_budget_template():
    """Create comprehensive budget calculator Excel template"""
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create all sheets
    create_project_summary_sheet(wb)
    create_monthly_budget_sheet(wb)
    create_channel_allocation_sheet(wb)
    create_production_tools_sheet(wb)
    create_weekly_tracking_sheet(wb)
    create_roi_calculator_sheet(wb)
    
    # Save the workbook
    filename = "AMP_Budget_Calculator_Template.xlsx"
    wb.save(filename)
    print(f"✅ Created: {filename}")
    return filename


def create_project_summary_sheet(wb):
    """Sheet 1: Project Summary"""
    ws = wb.create_sheet("📋 Project Summary")
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    label_font = Font(bold=True, size=11)
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "📋 งบประมาณโครงการ (Project Budget Summary)"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Project Information
    ws['A3'] = "ข้อมูลโครงการ (Project Information)"
    ws['A3'].font = label_font
    ws['A3'].fill = label_fill
    
    info_fields = [
        ("A5", "ชื่อโครงการ (Project Name)", "B5"),
        ("A6", "ผู้รับผิดชอบงบประมาณ (Budget Owner)", "B6"),
        ("A7", "ช่วงเวลา (Period)", "B7"),
        ("A8", "วันเริ่มแคมเปญ (Campaign Start)", "B8"),
        ("A9", "วันสิ้นสุดแคมเปญ (Campaign End)", "B9"),
        ("A10", "งบประมาณรวมที่อนุมัติ (Total Approved Budget)", "B10"),
    ]
    
    for label_cell, label_text, value_cell in info_fields:
        ws[label_cell] = label_text
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = ""
        ws[value_cell].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Budget Summary
    ws['A12'] = "สรุปงบประมาณรวม (Grand Total Summary)"
    ws['A12'].font = label_font
    ws['A12'].fill = label_fill
    
    ws['A14'] = "หมวดหมู่"
    ws['B14'] = "จำนวนเงิน (THB)"
    ws['C14'] = "% ของงบรวม"
    for cell in ['A14', 'B14', 'C14']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = label_fill
    
    categories = [
        "ค่าโฆษณา (Advertising Spend)",
        "ผลิตคอนเทนต์ (Content Production)",
        "เครื่องมือและซอฟต์แวร์ (Tools & Software)",
        "ค่าใช้จ่ายอื่นๆ (Other Expenses)",
    ]
    
    for idx, category in enumerate(categories, start=15):
        ws[f'A{idx}'] = category
        ws[f'B{idx}'] = 0
        ws[f'C{idx}'] = "0%"
    
    ws['A19'] = "งบประมาณรวมทั้งหมด (Grand Total)"
    ws['A19'].font = Font(bold=True, size=12)
    ws['B19'].font = Font(bold=True, size=12)
    ws['C19'].font = Font(bold=True, size=12)
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    apply_borders(ws, 'A3:C19')


def create_monthly_budget_sheet(wb):
    """Sheet 2: Monthly Budget Breakdown"""
    ws = wb.create_sheet("💵 Monthly Budget")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "💵 รายละเอียดงบประมาณรายเดือน (Monthly Budget Breakdown)"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Advertising Channels Header
    ws['A3'] = "ช่องทางโฆษณา (Advertising Channels)"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].fill = label_fill
    
    # Column headers
    headers = ["Channel/Campaign", "Budget (THB)", "% of Total", "Est. Clicks", "Est. CPL", "Est. Leads"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = label_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Channel data
    channels = [
        ("Google Ads", True),
        ("  - Search Campaigns", False),
        ("  - Display Network", False),
        ("  - Performance Max", False),
        ("Facebook/Instagram", True),
        ("  - Lead Gen Campaigns", False),
        ("  - Traffic Campaigns", False),
        ("  - Awareness Campaigns", False),
        ("LINE Ads", True),
        ("  - Talk Head View", False),
        ("  - Display Ads", False),
        ("TikTok Ads", True),
        ("  - In-Feed Ads", False),
        ("  - Spark Ads", False),
        ("YouTube Ads", True),
        ("Google Display Network", False),
        ("Other Channels", False),
    ]
    
    current_row = 5
    for channel, is_category in channels:
        ws[f'A{current_row}'] = channel
        if is_category:
            ws[f'A{current_row}'].font = Font(bold=True)
        else:
            ws[f'B{current_row}'] = 0
            ws[f'C{current_row}'] = "0%"
            ws[f'D{current_row}'] = 0
            ws[f'E{current_row}'] = 0
            ws[f'F{current_row}'] = 0
        current_row += 1
    
    # Total row
    total_row = current_row
    ws[f'A{total_row}'] = "รวมค่าโฆษณา (Total Ad Spend)"
    ws[f'A{total_row}'].font = Font(bold=True, size=11)
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{total_row}'].font = Font(bold=True)
        ws[f'{col}{total_row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    apply_borders(ws, f'A3:F{total_row}')


def create_channel_allocation_sheet(wb):
    """Sheet 3: Channel Budget Allocation"""
    ws = wb.create_sheet("🎯 Channel Allocation")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "🎯 การจัดสรรงบตามช่องทาง (Channel Budget Allocation)"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # By Marketing Funnel
    ws['A3'] = "การจัดสรรตาม Marketing Funnel"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].fill = label_fill
    
    headers = ["Funnel Stage", "Budget %", "Channels", "Objective"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = label_fill
    
    funnel_data = [
        ("Awareness", "20-30%", "TikTok, YouTube, FB/IG Video", "Reach & Brand awareness"),
        ("Consideration", "30-40%", "Google Display, FB/IG Traffic", "Interest & Engagement"),
        ("Conversion", "40-50%", "Google Search, FB/IG Leads", "Lead generation"),
    ]
    
    for idx, (stage, budget, channels, objective) in enumerate(funnel_data, start=5):
        ws[f'A{idx}'] = stage
        ws[f'B{idx}'] = budget
        ws[f'C{idx}'] = channels
        ws[f'D{idx}'] = objective
    
    # By Audience Type
    ws['A9'] = "การจัดสรรตามประเภทผู้ชม (Audience Type)"
    ws['A9'].font = Font(bold=True, size=11)
    ws['A9'].fill = label_fill
    
    headers2 = ["Audience Type", "Budget %", "Description", "Channels"]
    for idx, header in enumerate(headers2, start=1):
        cell = ws.cell(row=10, column=idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = label_fill
    
    audience_data = [
        ("Prospecting", "50-60%", "Cold audiences", "Lookalike, Interests, Keywords"),
        ("Retargeting", "30-40%", "Warm audiences", "Website visitors, Engagers"),
        ("Existing Leads", "5-10%", "Lead nurturing", "LINE OA, Email, Messenger"),
    ]
    
    for idx, (aud_type, budget, desc, channels) in enumerate(audience_data, start=11):
        ws[f'A{idx}'] = aud_type
        ws[f'B{idx}'] = budget
        ws[f'C{idx}'] = desc
        ws[f'D{idx}'] = channels
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 40
    
    apply_borders(ws, 'A3:D7')
    apply_borders(ws, 'A9:D13')


def create_production_tools_sheet(wb):
    """Sheet 4: Production & Tools Budget"""
    ws = wb.create_sheet("🎬 Production & Tools")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "🎬 งบผลิตและเครื่องมือ (Production & Tools Budget)"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = ["Category/Item", "Budget (THB)", "Frequency", "Notes"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = label_fill
    
    # Production items
    items = [
        ("Content Production", "", "", ""),
        ("  - Photography", 0, "One-time", "Professional property photos"),
        ("  - Videography", 0, "One-time", "Property tour video"),
        ("  - Drone Footage", 0, "One-time", "Aerial shots"),
        ("  - Graphics Design", 0, "Monthly", "Ad creatives & banners"),
        ("  - Copywriting", 0, "Monthly", "Ad copy & content"),
        ("", "", "", ""),
        ("Tools & Software", "", "", ""),
        ("  - Google Analytics", 0, "Free", "GA4, Search Console"),
        ("  - Canva Pro", 350, "Monthly", "Design tool"),
        ("  - Landing Page Builder", 0, "Monthly", "Unbounce/Leadpages"),
        ("  - CRM/Lead Management", 0, "Monthly", "Lead tracking system"),
        ("  - Ad Management Tool", 0, "Monthly", "Optional"),
        ("", "", "", ""),
        ("Other Expenses", "", "", ""),
        ("  - Agency Fee", 0, "Monthly", "If using agency"),
        ("  - Contingency (5-10%)", 0, "Monthly", "Buffer for unexpected costs"),
    ]
    
    current_row = 4
    for item, budget, frequency, notes in items:
        if item and not item.startswith("  -") and item != "":
            ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'] = item
        ws[f'B{current_row}'] = budget if budget != "" else ""
        ws[f'C{current_row}'] = frequency
        ws[f'D{current_row}'] = notes
        current_row += 1
    
    # Total
    ws[f'A{current_row}'] = "รวมงบผลิตและเครื่องมือ (Total)"
    ws[f'A{current_row}'].font = Font(bold=True, size=11)
    ws[f'B{current_row}'].font = Font(bold=True)
    ws[f'B{current_row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    
    apply_borders(ws, f'A3:D{current_row}')


def create_weekly_tracking_sheet(wb):
    """Sheet 5: Weekly Budget Tracking"""
    ws = wb.create_sheet("📅 Weekly Tracking")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "📅 ติดตามงบรายสัปดาห์ (Weekly Budget Tracking)"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = ["Week", "Ad Spend (THB)", "Clicks", "CTR (%)", "Leads", "CPL (THB)", "Budget Used %", "Pace"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = label_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Week rows
    for week in range(1, 5):
        row = 3 + week
        ws[f'A{row}'] = f"Week {week}"
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{row}'] = 0 if col in ['B', 'C', 'E', 'F'] else ""
    
    # Total row
    ws['A8'] = "Total"
    ws['A8'].font = Font(bold=True)
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}8'].font = Font(bold=True)
        ws[f'{col}8'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Budget pacing guide
    ws['A10'] = "Budget Pacing Guide:"
    ws['A10'].font = Font(bold=True, size=11)
    ws['A11'] = "On track: ~25% per week"
    ws['A12'] = "Under pace: < 20% per week"
    ws['A13'] = "Over pace: > 30% per week"
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15
    
    apply_borders(ws, 'A3:H8')


def create_roi_calculator_sheet(wb):
    """Sheet 6: ROI Calculator"""
    ws = wb.create_sheet("📊 ROI Calculator")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    value_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    result_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "📊 เครื่องคำนวณ ROI (ROI Calculator)"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Input Costs
    ws['A3'] = "ต้นทุนการลงทุน (Input Costs)"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].fill = label_fill
    
    costs = [
        ("Total Ad Spend (THB)", "B5"),
        ("Production Costs (THB)", "B6"),
        ("Agency/Management Fee (THB)", "B7"),
    ]
    
    row = 5
    for label, value_cell in costs:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[value_cell] = 0
        ws[value_cell].fill = value_fill
        row += 1
    
    ws['A8'] = "Total Investment (THB)"
    ws['A8'].font = Font(bold=True, size=11)
    ws['B8'].font = Font(bold=True)
    ws['B8'].fill = result_fill
    
    # Output Results
    ws['A10'] = "ผลลัพธ์ (Output Results)"
    ws['A10'].font = Font(bold=True, size=12)
    ws['A10'].fill = label_fill
    
    results = [
        ("Total Leads", "B12"),
        ("Qualified Leads", "B13"),
        ("Appointments", "B14"),
        ("Sales Closed", "B15"),
        ("Avg. Commission per Sale (THB)", "B16"),
    ]
    
    row = 12
    for label, value_cell in results:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[value_cell] = 0
        ws[value_cell].fill = value_fill
        row += 1
    
    ws['A17'] = "Total Revenue (THB)"
    ws['A17'].font = Font(bold=True, size=11)
    ws['B17'].font = Font(bold=True)
    ws['B17'].fill = result_fill
    
    # ROI Metrics
    ws['A19'] = "ตัวชี้วัด ROI (ROI Metrics)"
    ws['A19'].font = Font(bold=True, size=12)
    ws['A19'].fill = label_fill
    
    roi_metrics = [
        ("Net Profit (THB)", "B21"),
        ("ROI Percentage (%)", "B22"),
        ("ROI Ratio", "B23"),
        ("Break-even Sales", "B24"),
    ]
    
    row = 21
    for label, value_cell in roi_metrics:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[value_cell] = 0
        ws[value_cell].fill = result_fill
        ws[value_cell].font = Font(bold=True, color="0070C0")
        row += 1
    
    # ROI Formula
    ws['A26'] = "สูตรคำนวณ (Formula):"
    ws['A26'].font = Font(bold=True, size=11)
    ws['A27'] = "ROI % = (Revenue - Investment) / Investment × 100%"
    ws['A28'] = "ROI Ratio = Revenue / Investment"
    ws['A29'] = "Break-even = Total Investment / Commission per Sale"
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    
    apply_borders(ws, 'A3:B8')
    apply_borders(ws, 'A10:B17')
    apply_borders(ws, 'A19:B24')


def apply_borders(ws, cell_range):
    """Apply borders to a range of cells"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border


if __name__ == "__main__":
    print("🚀 Starting Budget Calculator Template Generation...")
    print("=" * 60)
    filename = create_budget_template()
    print("=" * 60)
    print(f"✅ Success! Template created: {filename}")
    print("\n📁 Template includes:")
    print("  1. 📋 Project Summary")
    print("  2. 💵 Monthly Budget")
    print("  3. 🎯 Channel Allocation")
    print("  4. 🎬 Production & Tools")
    print("  5. 📅 Weekly Tracking")
    print("  6. 📊 ROI Calculator")
    print("\n🎉 Ready to use for real estate advertising budget planning!")
